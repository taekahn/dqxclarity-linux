"""Sync curated community dialogue translations into the local cache.

The broad, human-translated dialogue corpus lives in the dqx-translation-project's
``dqx-custom-translations`` repo as ``merge.xlsx`` (sheets: Dialogue, Walkthrough, Quests,
Story So Far). Player names in it are stored as ``<pnplacehold>`` / ``<snplacehold>`` so a single
entry works for every player — see ``apply_name_placeholders`` in ``names``-aware lookup.

The community string API (community-string-api.ethene.wiki) is submit-only (GET → 405), so this
file-based snapshot is the way to get coverage without per-line MT.
"""

from __future__ import annotations

import io
import json
import re
import tarfile
import zipfile
from pathlib import Path

import httpx

from .db import TranslationCache

MERGE_XLSX_URL = (
    "https://github.com/dqx-translation-project/dqx-custom-translations/raw/main/csv/merge.xlsx"
)

# The dqx-custom-translations repo as a single codeload zip (default branch). Upstream
# common/constants.py uses this same archive URL (GITHUB_CUSTOM_TRANSLATIONS_ZIP_URL); we glob
# every json/*.json entry out of it rather than hand-picking a list — see download_custom_files()
# in dqxclarity/app/common/update.py:34-44, which imports EVERY member matching
# `"/json/" in name and name.endswith(".json")`.
CUSTOM_TRANSLATIONS_ZIP_URL = (
    "https://github.com/dqx-translation-project/dqx-custom-translations/archive/refs/heads/main.zip"
)
# Matches any ".../json/<file>.json" member in the repo archive (mirrors upstream's
# `"/json/" in obj.filename and obj.filename.endswith(".json")` predicate, update.py:38).
_CUSTOM_JSON_RE = re.compile(r"/json/.+\.json$")

# The ENTIRE human-translated static corpus lives in the dqx-translation-project's `dqx_translations`
# repo under json/_lang/en/ (~1000+ files: every quest, cutscene event, package, menu, name table).
# We pull the whole repo as one tarball and import every English JSON file, rather than a hand-picked
# list — so coverage is complete and stays current as the project adds files. All human-curated, so
# everything imports at the `community` quality tier.
TRANSLATIONS_TARBALL_URL = (
    "https://github.com/dqx-translation-project/dqx_translations/archive/refs/heads/main.tar.gz"
)
_EN_JSON_RE = re.compile(r"/json/_lang/en/[^/]+\.json$")

# The item-name JSON sources for the quest-reward dict (#21). Upstream maps these exact files to the
# m00_strings "files" used by clean_up_and_return_items ('items', 'key_items', 'custom_quest_rewards'):
#   * items     -> json/_lang/en/subPackage05Client.json       (constants.py:5)
#   * key_items -> json/_lang/en/subPackage41Client.win32.json (constants.py:6)
# from the dqx_translations tarball, and
#   * custom_quest_rewards -> json/custom_quest_rewards.json    (the dqx-custom-translations repo)
# from the dqx-custom-translations zip. We extract just these members (item name -> EN name) into a
# dict that is SEPARATE from the whole-string translation cache.
_ITEM_TARBALL_MEMBERS = (
    "json/_lang/en/subPackage05Client.json",
    "json/_lang/en/subPackage41Client.win32.json",
)
# Matches the custom_quest_rewards.json member anywhere under a json/ dir in the custom repo zip.
_CUSTOM_QUEST_REWARDS_RE = re.compile(r"/json/custom_quest_rewards\.json$")

# Sheets worth importing and the (japanese col, english col header substring) to read.
_SHEETS = ("Dialogue", "Walkthrough", "Quests", "Story So Far")

# Columns the BAD STRING markers live in. Upstream read_xlsx_and_import (update.py:134-154) reads the
# notes from column 4 and the "original bad string text" from column 5 (1-indexed). Our parser works
# off the header row, so we locate them by header substring and fall back to the upstream FIXED
# positions (0-indexed col 3 = notes, col 4 = original text) when no header matches.
_NOTES_FIXED_COL = 3            # upstream column 4 (1-indexed) -> 0-indexed 3
_ORIG_BAD_STRING_FIXED_COL = 4  # upstream column 5 (1-indexed) -> 0-indexed 4
# The marker substring upstream checks for in the notes column (update.py:148). A row is a BAD STRING
# row only when its notes cell CONTAINS this token (case-insensitive here for robustness).
_BAD_STRING_MARKER = "bad string"


def _notes_col(header: list) -> int | None:
    for i, h in enumerate(header):
        if "note" in str(h or "").lower():
            return i
    return _NOTES_FIXED_COL


def _orig_bad_string_col(header: list) -> int | None:
    for i, h in enumerate(header):
        h = str(h or "").lower()
        if "original" in h and ("bad" in h or "string" in h):
            return i
    return _ORIG_BAD_STRING_FIXED_COL


def _english_col(header: list) -> int | None:
    for i, h in enumerate(header):
        h = str(h or "").lower()
        if "english" in h and ("fixed" in h or "translation" in h):
            return i
    return None


def _deepl_col(header: list, en_col: int) -> int | None:
    """Locate the DeepL (machine-translation) column to fall back to for "Story So Far".

    Upstream (update.py:194-203) hard-codes DeepL at column 2 and fixed-en at column 3 — i.e. the
    DeepL column sits immediately before the fixed-english column. We prefer a header containing
    "deepl"; if none is found, fall back to the column right before fixed-english.
    """
    for i, h in enumerate(header):
        if "deepl" in str(h or "").lower():
            return i
    return en_col - 1 if en_col > 0 else None


def _cell(row: tuple, idx: int | None) -> str:
    if idx is None or idx < 0 or idx >= len(row) or not row[idx]:
        return ""
    val = str(row[idx])
    return "" if val == "None" else val


def parse_merge_xlsx(data: bytes) -> list[tuple[str, str]]:
    """Parse merge.xlsx bytes into (ja, en) rows with non-empty English.

    For every sheet we read the fixed-english column. For the "Story So Far" sheet specifically we
    fall back to the DeepL machine-translation column when fixed-english is empty, recovering the
    ~1,165 recaps that have only a DeepL rendering — see update.py:191-203, which uses fixed-en
    (col 3) and falls back to the DeepL column (col 2) when fixed-en is empty. Other sheets keep
    their fixed-english-only behavior.

    BAD STRING rows (notes column contains "BAD STRING") are SKIPPED here — they must NOT be imported
    on their (often partial/placeholder) primary ja key. They are handled by parse_merge_bad_strings
    instead: bad_string=1 -> a SUBSTRING suppression, bad_string=0 -> a FIX keyed on the column-5
    ORIGINAL (correct) ja. Importing such a row here would re-introduce the wrong-key bug the report
    flagged (a curated fix looked up by the partial string never hits).
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    rows: list[tuple[str, str]] = []
    for sheet in _SHEETS:
        if sheet not in wb.sheetnames:
            continue
        it = wb[sheet].iter_rows(values_only=True)
        header = list(next(it, ()) or ())
        en_col = _english_col(header)
        if en_col is None:
            continue
        notes_col = _notes_col(header)
        # Only "Story So Far" falls back to DeepL when fixed-english is empty.
        deepl_col = _deepl_col(header, en_col) if sheet == "Story So Far" else None
        for r in it:
            if not r or not r[0]:
                continue
            ja = str(r[0])
            # Skip BAD STRING rows; parse_merge_bad_strings owns them (suppression + correct-key fix).
            if _BAD_STRING_MARKER in _cell(r, notes_col).lower():
                continue
            en = _cell(r, en_col)
            if not en and deepl_col is not None:
                # DeepL sometimes echoes the JA back untranslated; drop those so a Story So Far
                # fallback never imports an untranslated cell (en == ja / placeholder markers).
                deepl = _cell(r, deepl_col)
                if not _is_untranslated(ja, deepl):
                    en = deepl
            if en:
                rows.append((ja, en))
    wb.close()
    return rows


def parse_merge_bad_strings(data: bytes) -> tuple[list[tuple[str, str]], list[tuple[str, str]]]:
    """Parse merge.xlsx into (suppressions, fixes) from its BAD STRING markers.

    Mirrors upstream read_xlsx_and_import (update.py:134-154), which marks a row as a BAD STRING
    when its notes column CONTAINS "BAD STRING" and then splits into two kinds:

      * bad_string=1 (SUPPRESSION): the notes say BAD STRING but there is NO column-5
        "original bad string text". Historically the primary ja here is a PARTIAL/placeholder
        string (player info redacted), so it can't be matched exactly — upstream's search_bad_strings
        scans every bad_strings row and returns en when ``ja in text`` (a SUBSTRING/contains match).
        We collect these as ``(ja_partial, en)`` suppression entries.

      * bad_string=0 (a normal curated FIX): the notes say BAD STRING AND a column-5 original text
        IS present. Upstream replaces ``source_text = original_bad_string_text`` and inserts it as an
        ordinary fix keyed on that CORRECT ja (not the partial). We collect these as ``(orig_ja, en)``
        fixes so they land in the normal cache keyed by the correct source (the report flagged
        keying the fix on the wrong/partial ja as the bug).

    Returns ``(suppressions, fixes)`` — both lists of ``(ja, en)``. ONLY the Dialogue sheet is
    scanned: upstream's read_xlsx_and_import (update.py:134-158) applies the notes/bad-string logic
    to the Dialogue sheet alone — Walkthrough/Quests/Story So Far have no Notes or original-bad-string
    columns and are imported as plain fixes there. Scanning them here would be overcoverage: our
    ``_notes_col`` falls back to the FIXED column 3 when a sheet has no Notes header, so any column-3
    data in a Walkthrough/Quests row that happened to contain "bad string" would be misclassified, and
    a non-Dialogue row could be mis-keyed on its column-4 (``orig_col``) data instead of the primary
    JA. If BAD STRING handling ever needs another sheet, add it here explicitly. A row needs a
    non-empty source AND english to be considered.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    suppressions: list[tuple[str, str]] = []
    fixes: list[tuple[str, str]] = []
    # Dialogue-only, per the upstream source of truth (update.py:134-158): the other sheets have no
    # Notes/original-bad-string columns, so scanning them for BAD STRING markers is overcoverage.
    for sheet in ("Dialogue",):
        if sheet not in wb.sheetnames:
            continue
        it = wb[sheet].iter_rows(values_only=True)
        header = list(next(it, ()) or ())
        en_col = _english_col(header)
        if en_col is None:
            continue
        notes_col = _notes_col(header)
        orig_col = _orig_bad_string_col(header)
        for r in it:
            if not r or not r[0]:
                continue
            ja = str(r[0])
            en = _cell(r, en_col)
            if not en:
                continue
            notes = _cell(r, notes_col)
            if _BAD_STRING_MARKER not in notes.lower():
                continue  # not a BAD STRING row -> handled by the normal parse_merge_xlsx path
            orig = _cell(r, orig_col)
            if orig:
                # Original (correct) source present -> a normal FIX keyed on the CORRECT ja.
                fixes.append((orig, en))
            else:
                # No original -> the primary ja is a partial/placeholder: a SUBSTRING suppression.
                suppressions.append((ja, en))
    wb.close()
    return suppressions, fixes


def sync_community(cache: TranslationCache, *, url: str = MERGE_XLSX_URL) -> int:
    """Download merge.xlsx and import its translations into the cache. Returns rows imported.

    Imports BOTH the ordinary fixed-english rows (parse_merge_xlsx) AND the bad_string=0 FIX rows
    (parse_merge_bad_strings), the latter keyed on the CORRECT ja (column-5 original text) so a
    curated fix is looked up by its real source rather than the partial/placeholder string. The
    bad_string=1 SUPPRESSION entries are NOT cached here (they aren't translations) — load them via
    ``load_suppressions`` and feed them to a SuppressionIndex on the dispatch path.
    """
    resp = httpx.get(url, timeout=60.0, follow_redirects=True)
    resp.raise_for_status()
    rows = parse_merge_xlsx(resp.content)
    _, fixes = parse_merge_bad_strings(resp.content)
    rows = rows + fixes
    cache.store_many([(ja, en, "community") for ja, en in rows])
    return len(rows)


def load_suppressions(*, url: str = MERGE_XLSX_URL) -> list[tuple[str, str]]:
    """Download merge.xlsx and return its bad_string=1 SUPPRESSION entries as ``(ja, en)``.

    These are the partial/placeholder strings whose machine translation broke or confused the game;
    upstream returns the curated ``en`` fallback (via a SUBSTRING match) instead of translating. Feed
    the result to ``translate.suppression.SuppressionIndex`` to install a pre-pass on the dispatch
    path. The fixes (bad_string=0) are imported into the cache by ``sync_community`` instead.

    This is the NETWORK fetch used by the ``sync`` command to refresh the local snapshot; ``run``
    must NOT call it. The fast-startup run path reads the persisted JSON via
    :func:`load_suppressions_local` instead, mirroring the download-once-at-sync / read-local-at-run
    model the TranslationCache and glossary already use.
    """
    resp = httpx.get(url, timeout=60.0, follow_redirects=True)
    resp.raise_for_status()
    suppressions, _ = parse_merge_bad_strings(resp.content)
    return suppressions


# Download alias (clearer name at the sync call site). The historical name ``load_suppressions`` is
# kept as the canonical download entry point so existing callers/tests keep working.
fetch_suppressions = load_suppressions


def save_suppressions(path: str | Path, entries: list[tuple[str, str]]) -> Path:
    """Persist the suppression entries as a JSON list-of-``[ja, en]`` pairs at ``path``.

    Used by ``sync`` to write the local snapshot the fast-startup ``run`` path reads back via
    :func:`load_suppressions_local`. Returns the written path. Tuples serialize as JSON arrays.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = [[ja, en] for ja, en in entries]
    path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    return path


def load_suppressions_local(path: str | Path) -> list[tuple[str, str]]:
    """Read the locally-persisted suppression entries written by :func:`save_suppressions`.

    LOCAL ONLY: never touches the network and never raises — a missing, unreadable, or malformed
    file degrades to ``[]`` so ``run`` builds an empty (no-op) SuppressionIndex and translation
    proceeds untouched. The fast-startup contract: ``run`` does ZERO network for this feature.
    """
    path = Path(path)
    try:
        obj = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return []
    if not isinstance(obj, list):
        return []
    out: list[tuple[str, str]] = []
    for item in obj:
        # Accept ``[ja, en]`` pairs (the saved shape); skip any malformed member rather than raise.
        if isinstance(item, (list, tuple)) and len(item) == 2:
            ja, en = item
            if isinstance(ja, str) and isinstance(en, str):
                out.append((ja, en))
    return out


def _is_untranslated(ja: str, en: str) -> bool:
    """True if ``en`` is not a usable English translation of ``ja`` and should be skipped.

    Drops empty/null English, English identical to the Japanese (still untranslated), and the
    project's placeholder markers for entries nobody has translated yet.
    """
    if not en:
        return True
    if en == ja:
        return True
    low = en.strip().lower()
    # Common "not yet translated" placeholders seen in the project's JSON sources.
    if low in ("", "null", "none", "untranslated", "no translation"):
        return True
    return False


def parse_translation_json(data: bytes) -> list[tuple[str, str]]:
    """Parse a dqx-translation-project static JSON file into (ja, en) rows.

    The sources are NOT all the same shape. Two are supported:

    * Nested — ``{"<id>": {"<japanese>": "<english>"}}``: each value is a single-entry dict
      mapping JA->EN (the quest file). We take that dict's key/value pair(s) as (ja, en).
    * Flat — ``{"<japanese>": "<english>"}``: the top-level value is a string, so the key is the
      Japanese and the value is the English.

    Entries whose English is empty/null, equal to the Japanese, or an untranslated placeholder are
    skipped. Unexpected shapes (non-dict top level, non-str/dict values, dicts with non-str
    members) are skipped rather than raised, so one malformed entry never aborts the import.
    """
    try:
        obj = json.loads(data)
    except (ValueError, TypeError):
        return []
    if not isinstance(obj, dict):
        return []

    rows: list[tuple[str, str]] = []
    for key, value in obj.items():
        if not isinstance(key, str):
            continue
        if isinstance(value, str):
            # Flat shape: key is JA, value is EN.
            ja, en = key, value
            if not _is_untranslated(ja, en):
                rows.append((ja, en))
        elif isinstance(value, dict):
            # Nested shape: each inner entry maps JA -> EN.
            for inner_ja, inner_en in value.items():
                if not isinstance(inner_ja, str) or not isinstance(inner_en, str):
                    continue
                if not _is_untranslated(inner_ja, inner_en):
                    rows.append((inner_ja, inner_en))
        # Any other value type (list, number, None, …) is an unexpected shape: skip it.
    return rows


def import_translation_tarball(content: bytes, cache: TranslationCache, *, batch: int = 20000) -> tuple[int, int]:
    """Import every json/_lang/en/*.json file from a dqx_translations repo tarball (gzip bytes).

    Returns (files_imported, rows_imported). Rows are stored in batches to bound memory. A file
    that fails to parse contributes 0 rows and is skipped (parse_translation_json never raises).
    """
    files = rows_total = 0
    buf: list[tuple[str, str, str]] = []
    with tarfile.open(fileobj=io.BytesIO(content), mode="r:gz") as tar:
        for member in tar:
            if not member.isfile() or not _EN_JSON_RE.search(member.name):
                continue
            extracted = tar.extractfile(member)
            if extracted is None:
                continue
            rows = parse_translation_json(extracted.read())
            if not rows:
                continue
            files += 1
            rows_total += len(rows)
            buf.extend((ja, en, "community") for ja, en in rows)
            if len(buf) >= batch:
                cache.store_many(buf)
                buf.clear()
    if buf:
        cache.store_many(buf)
    return files, rows_total


def sync_all_static(cache: TranslationCache, *, url: str = TRANSLATIONS_TARBALL_URL) -> tuple[int, int]:
    """Download the whole dqx_translations repo and import its entire English corpus."""
    resp = httpx.get(url, timeout=180.0, follow_redirects=True)
    resp.raise_for_status()
    return import_translation_tarball(resp.content, cache)


def import_custom_zip(content: bytes, cache: TranslationCache) -> tuple[int, int]:
    """Import EVERY ``*/json/*.json`` member from a dqx-custom-translations repo archive zip.

    Returns (files_imported, rows_imported). We glob the whole repo's json/ directory rather than
    hand-picking a list of filenames, so newly-added custom files (custom_corner_text,
    custom_npc_name_overrides, …) are pulled in automatically. Ported from
    dqxclarity/app/common/update.py:34-44, which imports every member matching
    `"/json/" in obj.filename and obj.filename.endswith(".json")`. A file that fails to parse
    contributes 0 rows and is skipped (parse_translation_json never raises).
    """
    files = rows_total = 0
    buf: list[tuple[str, str, str]] = []
    with zipfile.ZipFile(io.BytesIO(content)) as zf:
        for name in zf.namelist():
            if name.endswith("/") or not _CUSTOM_JSON_RE.search(name):
                continue
            rows = parse_translation_json(zf.read(name))
            if not rows:
                continue
            files += 1
            rows_total += len(rows)
            buf.extend((ja, en, "community") for ja, en in rows)
    if buf:
        cache.store_many(buf)
    return files, rows_total


def sync_custom_supplements(cache: TranslationCache, *, url: str = CUSTOM_TRANSLATIONS_ZIP_URL) -> int:
    """Download the dqx-custom-translations repo zip and import all of its json/*.json files.

    Returns rows imported. The whole repo archive is fetched once and every json member globbed in
    (see import_custom_zip) instead of fetching a hand-picked list file-by-file.
    """
    resp = httpx.get(url, timeout=180.0, follow_redirects=True)
    resp.raise_for_status()
    _, rows = import_custom_zip(resp.content, cache)
    return rows


# --------------------------------------------------------------------------------------------------
# #21 Quest-reward item dictionary (JA item name -> EN item name), kept SEPARATE from the cache.
# --------------------------------------------------------------------------------------------------


def _items_from_tarball(content: bytes) -> dict[str, str]:
    """Extract the item/key_item JA->EN dict from a dqx_translations tarball (gzip bytes).

    Reads ONLY the two reward-relevant members (subPackage05Client = items, subPackage41Client =
    key_items); each is the nested ``{id: {ja: en}}`` shape parse_translation_json already handles.
    A member that's missing or unparseable contributes nothing (never raises).
    """
    wanted = set(_ITEM_TARBALL_MEMBERS)
    out: dict[str, str] = {}
    with tarfile.open(fileobj=io.BytesIO(content), mode="r:gz") as tar:
        for member in tar:
            if not member.isfile():
                continue
            # Match by the member's path suffix (the archive prefixes a repo-name dir).
            if not any(member.name.endswith(w) for w in wanted):
                continue
            extracted = tar.extractfile(member)
            if extracted is None:
                continue
            for ja, en in parse_translation_json(extracted.read()):
                out[ja] = en
    return out


def _items_from_custom_zip(content: bytes) -> dict[str, str]:
    """Extract the custom_quest_rewards JA->EN dict from a dqx-custom-translations zip.

    Reads ONLY the ``json/custom_quest_rewards.json`` member (nested ``{id: {ja: en}}``). Missing or
    unparseable -> empty dict (never raises).
    """
    out: dict[str, str] = {}
    with zipfile.ZipFile(io.BytesIO(content)) as zf:
        for name in zf.namelist():
            if name.endswith("/") or not _CUSTOM_QUEST_REWARDS_RE.search(name):
                continue
            for ja, en in parse_translation_json(zf.read(name)):
                out[ja] = en
    return out


def build_reward_items_dict(
    *,
    tarball_url: str = TRANSLATIONS_TARBALL_URL,
    custom_url: str = CUSTOM_TRANSLATIONS_ZIP_URL,
) -> dict[str, str]:
    """Download + build the JA item-name -> EN item-name dict for quest-reward cleanup (#21).

    Merges items.json + key_items.json (from the dqx_translations tarball) with
    custom_quest_rewards.json (from the dqx-custom-translations zip), mirroring upstream's
    ``generate_m00_dict(files="'custom_quest_rewards', 'items', 'key_items'")`` (translate.py:508).
    The custom rewards are merged LAST so a curated reward name overrides the generic item name, the
    same precedence upstream's m00_strings INSERT ordering gives them. Returns a plain dict that
    ``rewards.clean_quest_rewards`` consumes; this dict is SEPARATE from the whole-string cache.
    """
    tar_resp = httpx.get(tarball_url, timeout=180.0, follow_redirects=True)
    tar_resp.raise_for_status()
    items = _items_from_tarball(tar_resp.content)

    custom_resp = httpx.get(custom_url, timeout=180.0, follow_redirects=True)
    custom_resp.raise_for_status()
    custom = _items_from_custom_zip(custom_resp.content)

    items.update(custom)  # custom_quest_rewards override generic item names
    return items


def load_reward_items(
    *,
    tarball_url: str = TRANSLATIONS_TARBALL_URL,
    custom_url: str = CUSTOM_TRANSLATIONS_ZIP_URL,
) -> dict[str, str]:
    """Loader alias for ``build_reward_items_dict`` — the quest-reward item-name dict for #21.

    This is the NETWORK build (downloads the dqx_translations tarball + custom zip). It is used by
    the ``sync`` command to refresh the local snapshot; ``run`` must NOT call it. The fast-startup
    run path reads the persisted JSON via :func:`load_reward_items_local` instead.
    """
    return build_reward_items_dict(tarball_url=tarball_url, custom_url=custom_url)


# Download alias (clearer name at the sync call site). The historical names
# ``build_reward_items_dict`` / ``load_reward_items`` remain the canonical download entry points so
# existing callers/tests keep working.
fetch_reward_items = load_reward_items


def save_reward_items(path: str | Path, items: dict[str, str]) -> Path:
    """Persist the JA item-name -> EN item-name dict as a JSON object at ``path``.

    Used by ``sync`` to write the local snapshot the fast-startup ``run`` path reads back via
    :func:`load_reward_items_local`. Returns the written path.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(items, ensure_ascii=False), encoding="utf-8")
    return path


def load_reward_items_local(path: str | Path) -> dict[str, str]:
    """Read the locally-persisted reward-item dict written by :func:`save_reward_items`.

    LOCAL ONLY: never touches the network and never raises — a missing, unreadable, or malformed
    file degrades to ``{}`` so ``run``'s quest reward fields fall back to the normal whole-string
    path. The fast-startup contract: ``run`` does ZERO network for this feature.
    """
    path = Path(path)
    try:
        obj = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return {}
    if not isinstance(obj, dict):
        return {}
    # Keep only str->str members; skip any malformed entry rather than raise.
    return {k: v for k, v in obj.items() if isinstance(k, str) and isinstance(v, str)}
